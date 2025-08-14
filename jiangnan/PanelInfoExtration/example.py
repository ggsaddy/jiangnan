import json
from io import StringIO
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment
import os
from dataclasses import dataclass
from typing import List
import openpyxl
from openpyxl.styles import Font, Alignment



def create_example_json_stream():
    """
    创建并返回一个样例JSON数据流
    返回:
        StringIO: 包含JSON数据的字符串流对象
    """
    # 创建StringIO流
    json_stream = StringIO()
    
    example_data = {
    "几何中心坐标": [-1281607.4416467769, 810361.8809703658],
    "边界轮廓": [
        {
            "类型": "边界",
            "编号": "边界1",
            "轮廓": ["line"],
            "句柄": ["2133F5"],
            "标注": [{
                "半径标注": {},
                "边长标注": {"句柄": "2133F5", "标注值": 300},
                "交点距离标注1": {},
                "交点距离标注2": {},
                "交点距离标注3": {},
                "交点距离标注4": {},
                "水平或竖直轴夹角标注": {},
                "定位标注": {}
            }],
            "是否有折边": "否"
        },
        {
            "类型": "角隅孔",
            "轮廓": ["line", "arc"],
            "句柄": ["2133F5", "2133F5"],
            "标注": [
                {
                    "半径标注": {},
                    "边长标注": {},
                    "交点距离标注1": {},
                    "交点距离标注2": {},
                    "交点距离标注3": {},
                    "交点距离标注4": {},
                    "水平或竖直轴夹角标注": {},
                    "定位标注": {}
                },
                {
                    "半径标注": {},
                    "边长标注": {},
                    "交点距离标注1": {},
                    "交点距离标注2": {},
                    "交点距离标注3": {},
                    "交点距离标注4": {},
                    "水平或竖直轴夹角标注": {},
                    "定位标注": {}
                }
            ]
        },
        {
            "类型": "边界",
            "编号": "边界2",
            "轮廓": ["line"],
            "句柄": ["2133F5"],
            "标注": [{
                "半径标注": {},
                "边长标注": {},
                "交点距离标注1": {
                    "句柄": "2133F5",
                    "标注值": 300,
                    "参考边": "边界3",
                    "交点坐标": ["x", "y"],
                    "顶点坐标": ["x", "y"],
                    "是否扩散": "否"
                },
                "交点距离标注2": {
                    "句柄": "2133F5",
                    "标注值": 300,
                    "参考边": "边界4",
                    "交点坐标": ["x", "y"],
                    "顶点坐标": ["x", "y"],
                    "是否扩散": "否"
                },
                "交点距离标注3": {},
                "交点距离标注4": {},
                "水平或竖直轴夹角标注": {},
                "定位标注": {}
            }],
            "是否有折边": "否"
        },
        {
            "类型": "角隅孔",
            "轮廓": ["line", "arc", "line"],
            "句柄": ["2133F5", "2133F5", "2133F5"],
            "标注": {"标注值": [20, 0, 55], "是否扩散": "否"},
            "文本标注": {"标注值": "20x12", "是否扩散": "否"}
        },
        {
            "句柄": "2133F5",
            "类型": "边界",
            "编号": "边界3",
            "轮廓": ["arc"],
            "句柄": ["2133F5"],
            "标注": [{
                "半径标注": {"句柄": "2133F5", "标注值": 50, "是否扩散": "否"},
                "边长标注": {},
                "交点距离标注1": {},
                "交点距离标注2": {},
                "交点距离标注3": {},
                "交点距离标注4": {},
                "水平或竖直轴夹角标注": {},
                "定位标注": {}
            }],
            "是否有折边": "否"
        },
        {
            "句柄": "2133F5",
            "类型": "边界",
            "编号": "边界4",
            "轮廓": ["arc"],
            "句柄": ["2133F5"],
            "标注": [{
                "半径标注": {},
                "边长标注": {},
                "交点距离标注1": {},
                "交点距离标注2": {},
                "交点距离标注3": {},
                "交点距离标注4": {},
                "水平或竖直轴夹角标注": {"句柄": "2133F5", "标注值": 30, "参考边": "水平轴", "是否扩散": "否"},
                "定位标注": {}
            }],
            "是否有折边": "否"
        },
        {
            "句柄": "2133F5",
            "类型": "边界",
            "编号": "边界5",
            "轮廓": ["arc"],
            "句柄": ["2133F5"],
            "标注": [{
                "半径标注": {},
                "边长标注": {},
                "交点距离标注1": {},
                "交点距离标注2": {},
                "交点距离标注3": {},
                "交点距离标注4": {},
                "水平或竖直轴夹角标注": {},
                "定位标注": {
                    "句柄": "2133F5",
                    "标注值": 30,
                    "参考边": "2133F5",
                    "参考边起点终点": ["x1", "x2", "y1", "y2"],
                    "是否扩散": "否"
                }
            }],
            "是否有折边": "否"
        }
    ],
    "多条标注边的标注信息列表": [
        {
            "平行距离标注": {
                "句柄": "2133F5",
                "标注值": 300,
                "标注边1": "2133F5",
                "标注边2": "2133F5",
                "是否扩散": "否"
            },
            "边界夹角标注": {
                "句柄": "2133F5",
                "标注值": 15,
                "标注边1": "2133F5",
                "标注边2": "2133F5",
                "是否扩散": "否"
            }
        }
    ]
}
    
    # 将数据写入json_stream
    json.dump(example_data, json_stream, ensure_ascii=False, indent=2)
    
    # 重置流指针到起始位置以便读取
    json_stream.seek(0)
    
    return json_stream



def save_json_to_file(json_stream, filepath):
    """
    将JSON流保存到当前目录下的output文件夹
    
    参数:
        json_stream: StringIO - 包含JSON数据的流对象
        filename: str - 输出文件名（默认output.json）
    
    返回:
        str - 生成的完整文件路径
    """
    # # 创建output目录（如果不存在）
    # output_dir = Path("output")
    # output_dir.mkdir(exist_ok=True)
    
    # # 构建完整文件路径
    # filepath = output_dir / filename
    
    try:
        # 从流中读取数据并直接写入文件
        with open(filepath, 'w', encoding='utf-8') as f:
            # 将流位置重置到开头以确保完整读取
            json_stream.seek(0)
            f.write(json_stream.read())
        
        print(f"JSON文件已保存到: {filepath.resolve()}")
        return str(filepath.resolve())
    
    except Exception as e:
        print(f"保存文件时出错: {str(e)}")
        raise
    


@dataclass
class PanelAnnotation:
    """肘板标注数据类"""
    poly_id: int
    annotation_handle: str
    geometric_center: str
    annotation_property: str
    annotated_edge: str
    edge1_handle: str
    edge2_handle: str
    reference_edge: str

class PanelAnno:
    """板材标注报告数据（包含完整表格数据）"""
    def __init__(self):
        self.annotations = self._load_complete_data()
    
    def _load_complete_data(self) -> List[PanelAnnotation]:
        """加载完整的表格数据"""
        return [
            PanelAnnotation(4, "E6EC", "(500054.6,320000.0)", "半径尺寸标注", "边界1", "6B95C", "6B95B", "8F876"),
            PanelAnnotation(6, "B95C", "(321716.0,280000.0)", "边长标注", "边界3", "1F894", "1F893", "无"),
            PanelAnnotation(8, "B95B", "(476837.2,350000.0)", "角度标注", "边界1", "DF2D", "1F36", "约束边4"),
            PanelAnnotation(8, "1F876", "(476837.2,350000.0)", "半径尺寸标注", "边界1", "1F894", "1F893", "无"),
            PanelAnnotation(8, "1F894", "(321716.0,280000.0)", "边长标注", "边界3", "1F893", "DF2D", "无"),
            PanelAnnotation(8, "1F893", "(321716.0,280000.0)", "半径尺寸标注", "边界1", "B95A", "28CE", "无"),
            PanelAnnotation(8, "DF2D", "(476837.2,350000.0)", "边长标注", "边界2", "1F36", "B95A", "无"),
            PanelAnnotation(8, "1F36", "(476837.2,350000.0)", "角度标注", "边界1", "B95A", "28CE", "无"),
            PanelAnnotation(8, "B95A", "(476837.2,350000.0)", "半径尺寸标注", "边界1", "28CE", "1F871", "无"),
            PanelAnnotation(8, "28CE", "(476837.2,350000.0)", "边长标注", "边界2", "1F871", "1F875", "无"),
            PanelAnnotation(8, "1F871", "(476837.2,350000.0)", "定位标注", "边界2", "1F875", "无", "无"),
            PanelAnnotation(8, "1F875", "(476837.2,350000.0)", "半径尺寸标注", "边界1", "无", "无", "约束边4")
        ]


    
def create_anno_excel(report_data: PanelAnno, output_path):
    """
    创建板材标注Excel文件
    
    参数:
        report_data: PanelAnno - 包含标注数据的报告对象
        output_file: str - 输出文件路径
    """
    # # 创建输出目录
    # Path("output").mkdir(exist_ok=True)
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "板材标注"
    
    # 设置表头
    headers = [
        "poly_id", "标注句柄", "板材几何中心", 
        "标注属性", "标注边",
        "标注边1句柄", "标注边2句柄", "参考边"
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 写入数据
    for row_num, annotation in enumerate(report_data.annotations, 2):
        row_data = [
            annotation.poly_id,
            annotation.annotation_handle,
            annotation.geometric_center,
            annotation.annotation_property,
            annotation.annotated_edge,
            annotation.edge1_handle,
            annotation.edge2_handle,
            annotation.reference_edge
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # 设置自动列宽
    for col in ws.columns:
        max_length = max(
            len(str(cell.value)) if cell.value else 0 
            for cell in col
        )
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width
    
    # 保存文件
    wb.save(output_path)
    print(f"标注表格已生成: {os.path.abspath(output_path)}")